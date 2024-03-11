import requests
from openpyxl import load_workbook
from openpyxl.styles import Font, Color
from openpyxl.styles.colors import RED, BLUE
from time import strftime


def set_color_for_text(cell, text):
    # Sets the cell color based on the test result.    
    if text in ["PASSED","Passed","Blue"]:
        cell.font = Font(color=Color(BLUE))
    elif text in ["FAILED","Failed","Red"]:
        cell.font = Font(color=Color(RED))

# Load the Excel file
wb = load_workbook('./TinyProjects/specs/Define.xlsx')
ws = wb.active

# Iterate through rows starting from the testing data row
for row in range(6, ws.max_row + 1):
    # Get target url
    e_value = ws.cell(row=1, column=5).value
    # Get column B and C values
    b_value = ws.cell(row=row, column=2).value
    c_value = ws.cell(row=row, column=3).value

    # Skip if B or C values is blank
    if not b_value or not c_value:
        ws.cell(row=row, column=5).value = f"Value is not defined"
        continue

    # Construct the URL
    url = f"https://{e_value}/{b_value}"

    # Verify the title and write test results to columns D and E
    try:
        # send a GET request
        response = requests.get(url)
        if response.status_code == 200:
            page_title = response.text.split('<title>')[1].split('</title>')[0]
            if c_value == page_title:
                ws.cell(row=row, column=4).value = page_title
                ws.cell(row=row, column=5).value = "Passed"
                set_color_for_text(ws.cell(row=row, column=5), "Passed")
            else:
                ws.cell(row=row, column=4).value = page_title
                ws.cell(row=row, column=5).value = "Failed"
                set_color_for_text(ws.cell(row=row, column=5), "Failed")
                ws.cell(row=row, column=6).value = f"Title '{page_title}' is not same as Expected Title."
        else:
            response.raise_for_status()
    except (requests.exceptions.HTTPError, 
            requests.exceptions.ConnectionError,
            requests.exceptions.Timeout,
            requests.exceptions.RequestException) as err:
        test_err = f"Connection Error: {err}"         
        ws.cell(row=1, column=7).value = test_err
        set_color_for_text(ws.cell(row=1, column=7), "FAILED")

# Save the workbook
wb.save(f"TestResult_{strftime('%Y%m%d_%H%M%S')}.xlsx")